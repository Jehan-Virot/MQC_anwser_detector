from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def clean_cell(value):
    """
    Convertit les None en cellules vides.
    Convertit les ID en str pour éviter les problèmes Excel.
    """
    if value is None:
        return ""
    return str(value)


def create_presence_workbook(xlsx_path):
    """
    Crée un fichier xlsx vide avec les 3 colonnes demandées.
    Si le fichier existe déjà, il est écrasé.
    """
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "PRESENCES"

    headers = ["imageName", "studentID_grid", "studentID_signature"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 24

    wb.save(xlsx_path)


def append_presence_row(xlsx_path, image_name, student_id_grid=None, student_id_signature=None,):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    ws.append([
        clean_cell(image_name),
        clean_cell(student_id_grid),
        clean_cell(student_id_signature),
    ])

    wb.save(xlsx_path)
