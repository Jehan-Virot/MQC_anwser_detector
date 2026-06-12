"""Write the two-sheet result workbook (PAGE-01 + EXAM) exactly like the
reference template produced by the teaching team.

PAGE-01: column A = label, column B = value (17 rows).
EXAM:    header row + one row per question; a marked choice is written as 1,
         numeric answers fill MANTISSE / EXPOSANT / UNITE.
"""

from datetime import datetime
import openpyxl
from openpyxl.styles import Font

from .layout_form3 import PAGE01_LABELS, EXAM_HEADER, EXAM_CHOICE_COLS

FONT = Font(name="Calibri", size=11)


def _date_value(s):
    """Convert a dd/mm/yyyy string to a datetime, else return the raw value."""
    if isinstance(s, datetime):
        return s
    if isinstance(s, str):
        try:
            return datetime.strptime(s.strip(), "%d/%m/%Y")
        except ValueError:
            return s
    return s


def write_workbook(out_path, page01_values, exam_answers, n_questions):
    """page01_values: dict label -> value (may be None).
    exam_answers: dict question_no -> {'choices': set('A'..'H'),
                                       'mantisse': float|None,
                                       'exposant': int|None,
                                       'unite': str|None}.
    """
    wb = openpyxl.Workbook()

    # ---- PAGE-01 ----
    ws = wb.active
    ws.title = "PAGE-01"
    for r, label in enumerate(PAGE01_LABELS, start=1):
        if label is None:
            continue  # preserve blank row in the template layout
        ws.cell(row=r, column=1, value=label).font = FONT
        val = page01_values.get(label)
        if label == "Date":
            val = _date_value(val)
        c = ws.cell(row=r, column=2, value=val)
        c.font = FONT
        if label == "Date" and isinstance(val, datetime):
            c.number_format = "yyyy-mm-dd"

    # ---- EXAM ----
    we = wb.create_sheet("EXAM")
    for c, head in enumerate(EXAM_HEADER, start=1):
        we.cell(row=1, column=c, value=head).font = FONT
    choice_col_index = {letter: EXAM_HEADER.index(f"CHOIX {letter}") + 1
                        for letter in [col.split()[-1] for col in EXAM_CHOICE_COLS]}
    mant_i = EXAM_HEADER.index("MANTISSE") + 1
    exp_i = EXAM_HEADER.index("EXPOSANT") + 1
    uni_i = EXAM_HEADER.index("UNITE") + 1

    for q in range(1, n_questions + 1):
        row = q + 1
        we.cell(row=row, column=1, value=q).font = FONT
        ans = exam_answers.get(q, {})
        for letter in ans.get("choices", set()):
            we.cell(row=row, column=choice_col_index[letter], value=1).font = FONT
        if ans.get("mantisse") is not None:
            we.cell(row=row, column=mant_i, value=ans["mantisse"]).font = FONT
        if ans.get("exposant") is not None:
            we.cell(row=row, column=exp_i, value=ans["exposant"]).font = FONT
        if ans.get("unite"):
            we.cell(row=row, column=uni_i, value=ans["unite"]).font = FONT

    wb.save(out_path)
    return out_path
