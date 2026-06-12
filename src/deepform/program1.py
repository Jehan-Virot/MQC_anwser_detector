"""PROGRAMME 1 - presence validation from first-page photos.

Produces EXAM_FORMXX_PRESENCES.xlsx with three columns:
    imageName, studentID_grid, studentID_signature

studentID_grid comes from the graphic OMR grid (low level).
studentID_signature comes from signature matching against
STUDENT_CLASS_SIGNATURES (needs the signature database / model).
"""

import os
from os.path import join, isdir

import openpyxl
from openpyxl.styles import Font

from . import io_utils, recognizers
from .registration import build_frame
from .omr_grid import read_digit_grid
from .layout_form3 import ROI

FONT = Font(name="Calibri", size=11)
IMG_EXT = (".jpg", ".jpeg", ".png", ".tif", ".tiff")


def autoValidID(image_path, signature_db_path):
    """Return (studentID_grid, studentID_signature) for one photo."""
    img = io_utils.load_image(image_path)
    frame, gray = build_frame(img)

    student_id_grid = None
    if frame is not None:
        sid_roi, _ = frame.roi(*ROI["student_id"])
        res = read_digit_grid(sid_roi, n_cols=5, n_rows=10)
        student_id_grid = res["value"]

    # signature -> student id (needs signature DB)
    student_id_sig = None
    if frame is not None:
        sig_roi, _ = frame.roi(*ROI["signature"])
        sid_sig, _, _ = recognizers.authenticate_signature(
            sig_roi, signature_db_path)
        student_id_sig = sid_sig

    return student_id_grid, student_id_sig


def autoValidPresences(presences_dir, signature_db_path, results_dir):
    """Write EXAM_FORMXX_PRESENCES.xlsx into results_dir."""
    os.makedirs(results_dir, exist_ok=True)
    out_path = join(results_dir, "EXAM_PRESENCES.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PRESENCES"
    for c, head in enumerate(["imageName", "studentID_grid",
                              "studentID_signature"], start=1):
        ws.cell(row=1, column=c, value=head).font = FONT

    if not isdir(presences_dir):
        wb.save(out_path)
        return out_path, []

    images = sorted(f for f in os.listdir(presences_dir)
                    if f.lower().endswith(IMG_EXT))
    rows = []
    for r, name in enumerate(images, start=2):
        sid_grid, sid_sig = autoValidID(join(presences_dir, name),
                                        signature_db_path)
        ws.cell(row=r, column=1, value=name).font = FONT
        ws.cell(row=r, column=2,
                value=int(sid_grid) if sid_grid else None).font = FONT
        ws.cell(row=r, column=3,
                value=int(sid_sig) if sid_sig else None).font = FONT
        rows.append((name, sid_grid, sid_sig))
        print(f"[autoValidPresences] {name}: grid={sid_grid} sig={sid_sig}")

    wb.save(out_path)
    return out_path, rows
